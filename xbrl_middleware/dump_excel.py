#! /usr/bin/env python
# encoding: utf-8

import xbrl
import xlsxwriter
import xlwt
import openpyxl


class ExcelDump(object):

    def __init__(self, xbrl_obj):
        self.xbrl_obj = xbrl_obj

    def __call__(self):
        return self.xbrl_obj


class XlsxWriterAdapter(ExcelDump):

    __dump = None
    __workbook = None

    def __init__(self, dump):
        self.__dump = dump

    def __call__(self):
        return self.__workbook

    def write(self, workbook_name='untitled.xlsx'):

        self.workbook_name = workbook_name
        self.__workbook = xlsxwriter.Workbook(self.workbook_name)
        worksheet = self.__workbook.add_worksheet()

        for idx, val in enumerate(self.__dump().__dict__.items()):
            worksheet.write(idx, 0, val[0].replace("_", " "))
            worksheet.write(idx, 1, val[1])
        # Explicit close() required for workbook, so no dump
        self.__workbook.close()


class XlwtAdapter(ExcelDump):

    __dump = None
    __workbook = None

    def __init__(self, dump):
        self.__dump = dump

    def __call__(self):
        return self.__workbook

    def write(self, workbook_name='untitled.xls', worksheet_name='untitled'):

        self.workbook_name = workbook_name
        self.__workbook = xlwt.Workbook()
        worksheet = self.__workbook.add_sheet(worksheet_name)

        for idx, val in enumerate(self.__dump().__dict__.items()):
            worksheet.write(idx, 0, val[0].replace("_", " "))
            worksheet.write(idx, 1, val[1])

    def dump(self):
        self.__workbook.save(self.workbook_name)


class OpenPyxlAdapter(ExcelDump):

    __dump = None
    __workbook = None

    def __init__(self, dump):
        self.__dump = dump

    def __call__(self):
        return self.__workbook

    def write(self, workbook_name='untitled.xlsx', worksheet_name='untitled'):

        self.workbook_name = workbook_name
        self.__workbook = openpyxl.Workbook(guess_types=True)
        worksheet = self.__workbook.active
        worksheet.title = worksheet_name

        for idx, val in enumerate(self.__dump().__dict__.items()):
            worksheet.cell(row=idx+1, column=1).value = val[0].replace("_", " ")
            worksheet.cell(row=idx+1, column=2).value = val[1]

    def dump(self):
        self.__workbook.save(self.workbook_name)
